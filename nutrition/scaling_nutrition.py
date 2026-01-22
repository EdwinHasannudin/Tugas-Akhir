import pandas as pd
from sklearn.preprocessing import MinMaxScaler


def apply_minmax_scaling(df, columns_to_scale=None):
    """
    Menerapkan Min-Max Scaling pada kolom nutrisi
    Min-Max Scaling: X_scaled = (X - X_min) / (X_max - X_min)
    Hasil akan bernilai antara 0-1
    
    Parameters:
    -----------
    df : pandas.DataFrame
        DataFrame yang akan di-scale
    columns_to_scale : list, optional
        Kolom yang akan di-scale. Default: ['calories', 'proteins', 'fat', 'carbohydrate']
    
    Returns:
    --------
    pandas.DataFrame
        DataFrame dengan kolom yang sudah di-scale
    """
    
    if columns_to_scale is None:
        columns_to_scale = ['calories', 'proteins', 'fat', 'carbohydrate']
    
    # Validasi kolom
    missing_cols = [col for col in columns_to_scale if col not in df.columns]
    if missing_cols:
        print(f"⚠️  Kolom tidak ditemukan: {missing_cols}")
        columns_to_scale = [col for col in columns_to_scale if col in df.columns]
    
    if not columns_to_scale:
        print("❌ Tidak ada kolom yang valid untuk di-scale")
        return df
    
    df_scaled = df.copy()
    
    # Terapkan MinMaxScaler
    scaler = MinMaxScaler()
    df_scaled[columns_to_scale] = scaler.fit_transform(df[columns_to_scale])
    
    return df_scaled


def save_to_excel(df, output_path="nutrition_scaled.xlsx"):
    """
    Menyimpan DataFrame ke Excel dengan format yang mudah dibaca
    
    Parameters:
    -----------
    df : pandas.DataFrame
        DataFrame yang akan disimpan
    output_path : str
        Path file Excel output
    """
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Nutrition Scaled', index=False)
        
        # Format Excel
        workbook = writer.book
        worksheet = writer.sheets['Nutrition Scaled']
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Data style
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Format data
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = border
                if cell.column > 1:  # Numerik columns
                    cell.alignment = center_alignment
                    if cell.column <= 5:  # Kolom nutrisi dengan 4 desimal
                        cell.number_format = '0.0000'
                else:  # Kolom name
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column width
        worksheet.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F']:
            worksheet.column_dimensions[col].width = 14
        if 'G' in worksheet.column_dimensions:
            worksheet.column_dimensions['G'].width = 16
    
    print(f"✓ File Excel tersimpan: {output_path}")


if __name__ == "__main__":
    # Baca file nutrisi
    input_file = "nutrition.csv"
    
    try:
        # Baca data
        df = pd.read_csv(input_file)
        print(f"📊 Total data: {len(df)} bahan makanan\n")
        
        # Terapkan scaling
        print("🔄 Menerapkan Min-Max Scaling...")
        df_scaled = apply_minmax_scaling(df)
        
        # Simpan ke Excel
        output_file = "nutrition_scaled.xlsx"
        save_to_excel(df_scaled, output_file)
        
        print("✅ Scaling selesai!\n")
        print("=" * 70)
        print("PREVIEW DATA HASIL SCALING (10 baris pertama)")
        print("=" * 70)
        preview = df_scaled[['name', 'calories', 'proteins', 'fat', 'carbohydrate']].head(10)
        print(preview.to_string(index=False))
        
    except FileNotFoundError:
        print(f"❌ File '{input_file}' tidak ditemukan")
    except Exception as e:
        print(f"❌ Error: {e}")
