"""
pembersihan.py - STEP 3 Pipeline
==================================
Membersihkan data hasil One-Hot Encoding:
  1. Hapus kategori 'lainnya'
  2. Hapus bahan berdasarkan kata kunci
  3. Strip kata tertentu dari nama bahan (misal: 'segar')
  4. Hapus duplikat (nama yang sama setelah strip)
  5. Simpan ke sheet 'Data Cleaned'

Alur Pipeline:
  scaling_nutrition.py (Step 1) -> onehot_encoding_nutrition.py (Step 2) -> pembersihan.py (Step 3)
"""

import re
import os
import pandas as pd
import warnings
warnings.filterwarnings('ignore')


# ============================================================================
# KONFIGURASI STRIP WORDS
# ============================================================================
# Kata-kata yang akan DIHILANGKAN dari nama bahan di kolom 'name'.
# Contoh: 'cumi-cumi segar' -> 'cumi-cumi', 'tomat merah segar' -> 'tomat merah'
# Tambahkan kata baru di list ini jika ingin menghilangkan kata lain dari nama bahan.
STRIP_WORDS = [
    'segar',
    # Tambahkan kata lain yang ingin dihilangkan, contoh:
    # 'lokal',
    # 'import',
]


# ============================================================================
# FUNGSI UTILITAS
# ============================================================================

def strip_words_from_name(name):
    """
    Bersihkan nama bahan dengan menghilangkan kata-kata di STRIP_WORDS.
    
    Contoh (jika STRIP_WORDS = ['segar']):
      'cumi-cumi segar'     -> 'cumi-cumi'
      'tomat merah segar'   -> 'tomat merah'
      'ayam kampung'        -> 'ayam kampung'  (tidak berubah)
    
    Cara kerja:
      - Setiap kata di STRIP_WORDS dicari sebagai whole-word (case-insensitive)
      - Kata yang cocok dihilangkan, lalu spasi berlebih dibersihkan
    
    Returns: (str_cleaned, list_of_removed_words)
    """
    if pd.isna(name):
        return name, []
    
    cleaned = str(name)
    removed = []
    
    for word in STRIP_WORDS:
        # Gunakan regex word boundary agar hanya hapus kata utuh
        # Contoh: 'segar' cocok di 'cumi-cumi segar' tapi tidak di 'menyegarkan'
        pattern = re.compile(r'\b' + re.escape(word) + r'\b', re.IGNORECASE)
        if pattern.search(cleaned):
            cleaned = pattern.sub('', cleaned)
            removed.append(word)
    
    # Bersihkan spasi berlebih
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    
    return cleaned, removed


def format_excel_sheet(worksheet, header_color="FF6B35"):
    """Format Excel sheet dengan styling profesional."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border
            try:
                if cell.value is not None:
                    float(cell.value)
                    cell.alignment = center_alignment
                    cell.number_format = '0.0000'
            except (ValueError, TypeError):
                cell.alignment = left_alignment
    
    for column in worksheet.columns:
        max_length = max((len(str(c.value or '')) for c in column), default=0)
        worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)


# ============================================================
# MAIN PROCESS
# ============================================================

print("=" * 80)
print("[STEP 3] PEMBERSIHAN DATA - STEP 3 (Pipeline)")
print("   Lanjutan dari: onehot_encoding_nutrition.py (Step 2)")
print("=" * 80)

# Baca file Excel dari STEP 2
input_file = 'dataset/nutrition_pipeline.xlsx'
print(f"\n[INFO] Membaca file: {input_file}")

if not os.path.exists(input_file):
    print(f"❌ File '{input_file}' tidak ditemukan!")
    print("   Pastikan Anda telah menjalankan onehot_encoding_nutrition.py terlebih dahulu")
    exit(1)

try:
    # Baca sheet dari Step 2 (One-Hot Encoded)
    df = pd.read_excel(input_file, sheet_name='One-Hot Encoded')
    print(f"   [OK] Membaca sheet 'One-Hot Encoded'")
except Exception:
    print("   ⚠️ Sheet 'One-Hot Encoded' tidak ditemukan. Mencari sheet alternatif...")
    try:
        df = pd.read_excel(input_file, sheet_name='Nutrition Scaled')
        print(f"   [OK] Membaca sheet 'Nutrition Scaled' (fallback)")
    except Exception:
        df = pd.read_excel(input_file, sheet_name=0)
        print(f"   [OK] Membaca sheet pertama (fallback)")

df_original = df.copy()
print(f"   ✓ Total data awal: {len(df)} bahan makanan\n")

# ============================================================
# Step 1: Hapus kategori 'lainnya'
# ============================================================
print("[STEP] 1: Menghapus kategori 'lainnya'...")
if 'kategori' in df.columns:
    df_cleaned = df[df['kategori'] != 'lainnya']
    removed_lainnya = len(df) - len(df_cleaned)
    print(f"   - Baris yang dihapus: {removed_lainnya}")
    print(f"   - Total setelah: {len(df_cleaned)}\n")
else:
    df_cleaned = df.copy()
    removed_lainnya = 0
    print("   ⚠️ Kolom 'kategori' tidak ditemukan, skip step ini\n")

# ============================================================
# Step 2: Hapus bahan berdasarkan kata kunci
# ============================================================
print("[STEP] 2: Menghapus bahan makanan berdasarkan kata kunci...")
kata_kunci_dihapus = [
    'andaliman', 'pohon', 'babi', 'anjing', 'kuda', 'domba', 'burung', 'angsa', 
    'belut', 'katak', 'keong', 'kodok', 'kura-kura', 'ulat sagu', 'ham', 'hiu', 
    'kotiu', 'buaya', 'dodol', 'penyu', 'masakan', 'goreng', 'kukus', 'lodeh',
    'ginjal', 'sosis', 'lilin', 'pempek', 'kue', 'martabak', 'otak', 'ular', 
    'purundawa', 'putri malu', 'purut', 'sate', 'sarimuka', 'tekwan', 'tinira', 
    'camilan', 'ketoprak', 'rusa', 'soto', 'anak', 'dideh/darah', 'hati', 'alabio', 
    'belibis', 'asap', 'dendeng', 'dideh', 'kering', 'bakar', 'jantung', 'perut',
    'mentah', 'hitam', 'kacangan', 'kawalinya', 'kima', 'lidah', 'nasu', 
    'betok', 'telan', 'bubuk', 'tepung', 'kerupuk', 'abon', 'gemuk', 'kornet',
    'kurus', 'lemak', 'kambing daging', 'kerbau daging', 'akar', 'asinan', 
    'rebus', 'sop', 'umbut', 'keripik', 'tondano', 'bader', 'balong', 'bambangan', 
    'baung', 'bekasang', 'belida', 'beunteur', 'biawan', 'bili', 'bubara', 'bulan-bulan', 
    'kakatua', 'kapar', 'katombo', 'layur', 'lehoma', 'lemuru', 'malalugis', 'pepes', 
    'mayong', 'oci', 'pepetek', 'sale', 'saluang', 'selar', 'sepat', 'sidat', 'sunu',
    'tahuman', 'tarmon', 'tembang', 'tempahas', 'terbang', 'titang', 'turi', 'peda', 'petis',
    'pisang', 'pindang', 'daleman', 'keleponan', 'sapi usus', 'tumis', 'sayur', 'belimbing',
    'ketupat', 'koro', 'lamtoro', 'tahu telur',  'ceplok', 'dadar', 'kampung', 'usus', 'liver',
    'ceplok', 'asin', 'terubuk', 'bongkrek', 'gembus', 'kacang', 'makanan', 'asam', 'madura',
    'hintalo', 'ampas', 'mie', 'kelinci', 
]

before_keyword = len(df_cleaned)
for kata in kata_kunci_dihapus:
    if 'name' in df_cleaned.columns:
        df_cleaned = df_cleaned[~df_cleaned['name'].str.contains(kata, case=False, na=False)]

removed_keywords = before_keyword - len(df_cleaned)
print(f"   - Baris yang dihapus: {removed_keywords}")
print(f"   - Total setelah: {len(df_cleaned)}\n")

# ============================================================
# Step 3: Strip kata tertentu dari nama bahan (STRIP_WORDS)
# ============================================================
print("[STEP] 3: Menghilangkan kata dari nama bahan (strip words)...")
print(f"   Kata yang dihilangkan: {STRIP_WORDS}")
stripped_count = 0

if 'name' in df_cleaned.columns:
    new_names = []
    for idx, row in df_cleaned.iterrows():
        original_name = row['name']
        cleaned_name, removed_words = strip_words_from_name(original_name)
        if removed_words:
            print(f"     [STRIP] '{original_name}' -> '{cleaned_name}' (dihapus: {', '.join(removed_words)})")
            stripped_count += 1
        new_names.append(cleaned_name)
    
    df_cleaned = df_cleaned.copy()
    df_cleaned['name'] = new_names
    print(f"   - Nama yang diubah: {stripped_count}")
    print(f"   - Total data: {len(df_cleaned)}\n")
else:
    print("   ⚠️ Kolom 'name' tidak ditemukan, skip step ini\n")

# ============================================================
# Step 4: Hapus duplikat (nama sama setelah strip)
# ============================================================
print("[STEP] 4: Menghapus data duplikat (nama yang sama)...")
if 'name' in df_cleaned.columns:
    # Buat kolom bantu untuk perbandingan case-insensitive
    df_cleaned['_name_lower'] = df_cleaned['name'].str.lower().str.strip()
    
    # Cari dan tampilkan duplikat
    duplicated_mask = df_cleaned.duplicated(subset='_name_lower', keep='first')
    duplicated_names = df_cleaned.loc[duplicated_mask, 'name'].tolist()
    
    if duplicated_names:
        print(f"   - Duplikat ditemukan ({len(duplicated_names)}):")
        for dup_name in duplicated_names:
            print(f"     x {dup_name}")
    else:
        print("   - Tidak ada duplikat ditemukan")
    
    # Hapus duplikat, simpan yang pertama
    before_dedup = len(df_cleaned)
    df_cleaned = df_cleaned.drop_duplicates(subset='_name_lower', keep='first')
    removed_duplicates = before_dedup - len(df_cleaned)
    
    # Hapus kolom bantu
    df_cleaned = df_cleaned.drop(columns=['_name_lower'])
    
    print(f"   - Baris duplikat dihapus: {removed_duplicates}")
    print(f"   - Total setelah: {len(df_cleaned)}\n")
else:
    removed_duplicates = 0
    print("   ⚠️ Kolom 'name' tidak ditemukan, skip step ini\n")

# ============================================================
# Step 5: Simpan hasil ke sheet 'Data Cleaned'
# ============================================================
print(f"[WRITE] Menambahkan sheet 'Data Cleaned' ke {input_file}...")

from openpyxl import load_workbook

with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_cleaned.to_excel(writer, sheet_name='Data Cleaned', index=False)
    format_excel_sheet(writer.sheets['Data Cleaned'], header_color="FF6B35")

print("[OK] Sheet 'Data Cleaned' berhasil ditambahkan!\n")

# ============================================================
# RINGKASAN
# ============================================================
print("=" * 80)
print("[RESULT] RINGKASAN PEMBERSIHAN DATA")
print("=" * 80)
print(f"\nData Awal       : {len(df_original)} bahan makanan")
print(f"Hapus 'lainnya' : -{removed_lainnya} baris")
print(f"Hapus keyword   : -{removed_keywords} baris")
print(f"Strip words     : {stripped_count} nama diubah")
print(f"Hapus duplikat  : -{removed_duplicates} baris")
print(f"Data Akhir      : {len(df_cleaned)} bahan makanan")
print(f"Total dihapus   : {len(df_original) - len(df_cleaned)} baris ({((len(df_original) - len(df_cleaned))/len(df_original)*100):.1f}%)")

if 'kategori' in df_cleaned.columns:
    print(f"\n[KATEGORI] Kategori yang tersisa:")
    kategori_counts = df_cleaned['kategori'].value_counts()
    for cat, count in kategori_counts.items():
        print(f"   - {cat}: {count} items")

print("\n[DONE] STEP 3 SELESAI: Pembersihan data berhasil disimpan!")
print(f"📁 File: {input_file}")
print("\n" + "=" * 80)
print("NEXT STEP: Jalankan pembagian.py")
print("=" * 80)