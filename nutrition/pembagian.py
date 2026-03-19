"""
pembagian.py - STEP 4 Pipeline
==================================
Membagi data yang sudah dibersihkan berdasarkan kategori.
Hanya mengambil kategori 'lauk' dan 'sayuran', lalu menyimpan
masing-masing sebagai sheet terpisah di nutrition_pipeline.xlsx.

Alur Pipeline:
  scaling_nutrition.py (Step 1) 
  -> onehot_encoding_nutrition.py (Step 2) 
  -> pembersihan.py (Step 3) 
  -> pembagian.py (Step 4) ← file ini
"""

import os
import pandas as pd
import warnings
warnings.filterwarnings('ignore')


# ============================================================================
# KONFIGURASI
# ============================================================================

# Kategori yang ingin diambil dan dijadikan sheet terpisah
TARGET_CATEGORIES = ['lauk', 'sayuran']


# ============================================================================
# FUNGSI UTILITAS
# ============================================================================

def format_excel_sheet(worksheet, header_color="27AE60"):
    """Format Excel sheet dengan styling profesional."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border
            try:
                if cell.value is not None:
                    float(cell.value)
                    cell.alignment = center
                    cell.number_format = '0.0000'
            except (ValueError, TypeError):
                cell.alignment = left

    for column in worksheet.columns:
        max_length = max((len(str(c.value or '')) for c in column), default=0)
        worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)


# ============================================================
# MAIN PROCESS
# ============================================================

print("=" * 80)
print("[STEP 4] PEMBAGIAN DATA PER KATEGORI - STEP 4 (Pipeline)")
print("   Lanjutan dari: pembersihan.py (Step 3)")
print("=" * 80)

# Baca file Excel dari STEP 3
input_file = 'dataset/nutrition_pipeline.xlsx'
print(f"\n[INFO] Membaca file: {input_file}")

if not os.path.exists(input_file):
    print(f"❌ File '{input_file}' tidak ditemukan!")
    print("   Pastikan Anda telah menjalankan pembersihan.py terlebih dahulu")
    exit(1)

try:
    # Baca sheet dari Step 3 (Data Cleaned)
    df = pd.read_excel(input_file, sheet_name='Data Cleaned')
    print(f"   [OK] Membaca sheet 'Data Cleaned'")
except Exception:
    print("   ❌ Sheet 'Data Cleaned' tidak ditemukan!")
    print("   Pastikan Anda telah menjalankan pembersihan.py terlebih dahulu")
    exit(1)

print(f"   ✓ Total data: {len(df)} bahan makanan")

# Cek kolom kategori
if 'kategori' not in df.columns:
    print("   ❌ Kolom 'kategori' tidak ditemukan!")
    exit(1)

# Tampilkan semua kategori yang tersedia
all_categories = df['kategori'].unique()
print(f"\n[INFO] Kategori yang tersedia: {sorted(all_categories)}")
print(f"[INFO] Kategori yang akan diambil: {TARGET_CATEGORIES}\n")

# ============================================================
# Proses: Filter dan simpan per kategori
# ============================================================

print("[STEP] Membagi data per kategori...")

with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    for kategori in TARGET_CATEGORIES:
        # Filter data untuk kategori ini (case-insensitive)
        df_kategori = df[df['kategori'].str.lower() == kategori.lower()]
        
        if len(df_kategori) == 0:
            print(f"   ⚠️ Kategori '{kategori}' tidak ditemukan dalam data, skip")
            continue
        
        # Nama sheet = nama kategori
        sheet_name = kategori
        
        # Tulis ke sheet
        df_kategori.to_excel(writer, sheet_name=sheet_name, index=False)
        format_excel_sheet(writer.sheets[sheet_name])
        
        print(f"   ✅ Sheet '{sheet_name}' dibuat: {len(df_kategori)} bahan makanan")

# ============================================================
# RINGKASAN
# ============================================================
print("\n" + "=" * 80)
print("[RESULT] RINGKASAN PEMBAGIAN DATA")
print("=" * 80)
print(f"\nFile: {input_file}")
print(f"Data dari sheet: 'Data Cleaned' ({len(df)} bahan)")

total_exported = 0
for kategori in TARGET_CATEGORIES:
    df_kat = df[df['kategori'].str.lower() == kategori.lower()]
    count = len(df_kat)
    total_exported += count
    print(f"   -> Sheet '{kategori}': {count} bahan makanan")

print(f"\nTotal data diekspor: {total_exported} bahan makanan")
print(f"Kategori lain (tidak diekspor): {len(df) - total_exported} bahan makanan")

print("\n[DONE] STEP 4 SELESAI: Pembagian data berhasil disimpan!")
print(f"📁 File: {input_file}")
print("\n" + "=" * 80)
print("PIPELINE SELESAI!")
print("Sheet yang tersedia di nutrition_pipeline.xlsx:")
print("   1. 'Dataset Asli'        - Data original dari CSV")
print("   2. 'Nutrition Scaled'    - Data setelah Min-Max Scaling")
print("   3. 'One-Hot Encoded'     - Data setelah One-Hot Encoding")
print("   4. 'Data Cleaned'        - Data setelah pembersihan")
print("   5. 'lauk'                - Data kategori lauk")
print("   6. 'sayuran'             - Data kategori sayuran")
print("=" * 80)