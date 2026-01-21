import pandas as pd

# Membaca data CSV
df = pd.read_csv('nutrition.csv')

# Membuat Excel writer object
with pd.ExcelWriter('nutrition_data.xlsx', engine='openpyxl') as writer:
    
    # Sheet 1: Data Utama
    df.to_excel(writer, sheet_name='Data Nutrition', index=False)
    
    # Sheet 2: Ringkasan Statistik
    summary_data = {
        'Metric': ['Rata-rata', 'Maksimum', 'Minimum', 'Total Item'],
        'Kalori': [df['calories'].mean(), df['calories'].max(), df['calories'].min(), len(df)],
        'Protein': [df['proteins'].mean(), df['proteins'].max(), df['proteins'].min(), ''],
        'Lemak': [df['fat'].mean(), df['fat'].max(), df['fat'].min(), ''],
        'Karbohidrat': [df['carbohydrate'].mean(), df['carbohydrate'].max(), df['carbohydrate'].min(), '']
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Ringkasan Statistik', index=False)
    
    # Sheet 3: Top 10 Makanan Tertinggi Kalori
    top_calories = df.nlargest(10, 'calories')[['name', 'calories', 'proteins', 'fat', 'carbohydrate']]
    top_calories.to_excel(writer, sheet_name='Top 10 Kalori', index=False)
    
    # Sheet 4: Top 10 Makanan Tertinggi Protein
    top_protein = df.nlargest(10, 'proteins')[['name', 'proteins', 'calories', 'fat', 'carbohydrate']]
    top_protein.to_excel(writer, sheet_name='Top 10 Protein', index=False)

# Formatting Excel (opsional - memerlukan openpyxl)
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Memuat workbook yang sudah dibuat
    wb = load_workbook('nutrition_data.xlsx')
    
    # Format untuk setiap sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Simpan perubahan formatting
    wb.save('nutrition_data.xlsx')
    print("✅ File Excel berhasil dibuat dengan formatting!")
    
except ImportError:
    print("✅ File Excel berhasil dibuat! (tanpa formatting tambahan)")

print(f"📊 Total data: {len(df)} makanan")
print(f"🔥 Kalori tertinggi: {df['calories'].max()} kcal - {df.loc[df['calories'].idxmax(), 'name']}")
print(f"💪 Protein tertinggi: {df['proteins'].max()}g - {df.loc[df['proteins'].idxmax(), 'name']}")